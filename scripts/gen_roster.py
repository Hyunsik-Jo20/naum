# 더미 학생 명부 생성 → 엑셀(로컬 운영 자료) + 앱용 roster.ts
# 규칙: 초등 6개 학년, 학년별 2~3개 반, 반당 20~25명
#       남학생 번호 1~19, 여학생 번호 30~50(30번부터 시작)
import os
import random
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

random.seed(42)

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

SURNAMES = list("김이박최정강조윤장임한오서신권황안송류홍")
MALE = ["민준","서준","도윤","예준","시우","하준","주원","지호","지후","준서",
        "준우","현우","도현","건우","우진","선우","서진","연우","정우","승우"]
FEMALE = ["서연","서윤","지우","서현","하은","하윤","민서","지유","윤서","채원",
          "수아","지아","지윤","은서","다은","예은","수빈","소율","예린","지율"]

# 학년별 반 수 (2 또는 3)
CLASSES_PER_GRADE = {1: 3, 2: 3, 3: 2, 4: 3, 5: 2, 6: 2}

rows = []  # (grade, cls, number, name, sex, sid)
for grade in range(1, 7):
    for cls in range(1, CLASSES_PER_GRADE[grade] + 1):
        boys = random.randint(9, 13)      # 번호 1..boys (<=19)
        girls = random.randint(10, 13)    # 번호 30..30+girls-1 (<=50)
        # 반당 20~25명 보정
        total = boys + girls
        while total < 20:
            girls += 1; total = boys + girls
        while total > 25:
            if girls > boys: girls -= 1
            else: boys -= 1
            total = boys + girls
        for n in range(1, boys + 1):
            name = random.choice(SURNAMES) + random.choice(MALE)
            rows.append((grade, cls, n, name, "남", f"s{grade}_{cls}_{n}"))
        for i in range(girls):
            num = 30 + i
            name = random.choice(SURNAMES) + random.choice(FEMALE)
            rows.append((grade, cls, num, name, "여", f"s{grade}_{cls}_{num}"))

# --- 엑셀 ---
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "학생명부"
headers = ["학년", "반", "번호", "이름", "성별", "학생ID"]
ws.append(headers)
head_font = Font(bold=True, color="FFFFFF")
head_fill = PatternFill("solid", fgColor="185FA5")
for c in range(1, len(headers) + 1):
    cell = ws.cell(1, c)
    cell.font = head_font
    cell.fill = head_fill
    cell.alignment = Alignment(horizontal="center")
for r in rows:
    ws.append(list(r))
widths = [8, 6, 8, 14, 8, 12]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[chr(64 + i)].width = w
ws.freeze_panes = "A2"

os.makedirs(os.path.join(BASE, "local-data"), exist_ok=True)
xlsx_path = os.path.join(BASE, "local-data", "students.xlsx")
wb.save(xlsx_path)

# --- roster.ts (앱 로컬 데이터) ---
ts_lines = [
    "// 자동 생성됨 (scripts/gen_roster.py). 학교 로컬에만 존재하는 학생 명부(PII).",
    "import type { Student } from '../types'",
    "",
    "export const roster: Student[] = [",
]
for (grade, cls, number, name, sex, sid) in rows:
    ts_lines.append(
        f"  {{ id: '{sid}', name: '{name}', grade: {grade}, classNo: {cls}, number: {number}, sex: '{sex}' }},"
    )
ts_lines.append("]")
ts_lines.append("")
ts_path = os.path.join(BASE, "src", "data", "roster.ts")
with open(ts_path, "w", encoding="utf-8") as f:
    f.write("\n".join(ts_lines))

# 요약
from collections import Counter
per_class = Counter((g, c) for (g, c, *_ ) in rows)
print(f"총 학생 {len(rows)}명 / 학급 {len(per_class)}개")
print("학급별 인원:", dict(sorted(per_class.items())))
print("xlsx:", xlsx_path)
print("ts  :", ts_path)
